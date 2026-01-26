"""
Sistema de Geração de Relatórios e Exportação
Suporta PDF e Excel para estatísticas pessoais e comparativas
"""
import io
from typing import Dict, List, Optional, Any
from datetime import datetime
from pathlib import Path

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.utils.database import IARTESDatabase


class ReportGenerator:
    """Gera relatórios em PDF e Excel"""
    
    def __init__(self, db: IARTESDatabase):
        self.db = db
    
    def generate_personal_stats_pdf(self, user_id: int) -> bytes:
        """
        Gera relatório PDF com estatísticas pessoais do usuário.
        
        Args:
            user_id: ID do usuário
        
        Returns:
            Bytes do arquivo PDF
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab não está instalado. Execute: pip install reportlab")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Buscar dados do usuário
        user = self.db.get_user_by_id(user_id)
        if not user:
            raise ValueError(f"Usuário {user_id} não encontrado")
        
        # Buscar estatísticas do usuário
        cursor = self.db.conn.cursor()
        cursor.execute("""
            SELECT * FROM user_learning_stats WHERE user_id = ?
        """, (user_id,))
        row = cursor.fetchone()
        user_stats = dict(row) if row else None
        
        # Título
        story.append(Paragraph("Relatório de Estatísticas Pessoais", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Informações do usuário
        story.append(Paragraph(f"<b>Usuário:</b> {user['full_name'] or user['username']}", styles['Normal']))
        story.append(Paragraph(f"<b>Nível de Experiência:</b> {user.get('experience_level', 'N/A')}", styles['Normal']))
        story.append(Paragraph(f"<b>Data do Relatório:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Estatísticas principais
        if user_stats:
            stats_data = [
                ['Métrica', 'Valor'],
                ['Total de Feedbacks', str(user_stats.get('total_feedbacks', 0))],
                ['Taxa de Sucesso', f"{user_stats.get('success_rate', 0) * 100:.1f}%"],
                ['Avaliação Média', f"{user_stats.get('avg_rating', 0):.2f}/5.0"],
                ['Tempo Médio de Execução', f"{user_stats.get('avg_execution_time', 0):.1f}s"],
            ]
            
            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Módulos preferidos
            preferred_modules = user_stats.get('preferred_modules', '')
            if preferred_modules:
                modules = preferred_modules.split(',')
                story.append(Paragraph("<b>Módulos Preferidos:</b>", styles['Heading2']))
                story.append(Paragraph(", ".join(modules), styles['Normal']))
                story.append(Spacer(1, 0.2*inch))
        
        # Informações do modelo personalizado
        cursor = self.db.conn.cursor()
        cursor.execute("""
            SELECT training_samples, last_trained FROM user_models WHERE user_id = ?
        """, (user_id,))
        model_row = cursor.fetchone()
        
        if model_row:
            story.append(Paragraph("<b>Modelo Personalizado:</b>", styles['Heading2']))
            story.append(Paragraph(f"Amostras de Treinamento: {model_row['training_samples'] or 0}", styles['Normal']))
            if model_row['last_trained']:
                story.append(Paragraph(f"Último Treinamento: {model_row['last_trained']}", styles['Normal']))
        
        # Histórico de feedbacks recentes
        story.append(PageBreak())
        story.append(Paragraph("Histórico de Feedbacks Recentes", styles['Heading1']))
        story.append(Spacer(1, 0.2*inch))
        
        feedbacks = self.db.get_user_feedbacks(user_id, limit=20)
        if feedbacks:
            feedback_data = [['Data', 'Teste', 'Sucesso', 'Avaliação', 'Tempo (s)']]
            for fb in feedbacks[:20]:
                feedback_data.append([
                    fb['executed_at'][:10],
                    fb['test_case_id'],
                    'Sim' if fb['success'] else 'Não',
                    str(fb['tester_rating']) if fb['tester_rating'] else 'N/A',
                    f"{fb['actual_execution_time']:.1f}"
                ])
            
            feedback_table = Table(feedback_data, colWidths=[1*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.8*inch])
            feedback_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            story.append(feedback_table)
        else:
            story.append(Paragraph("Nenhum feedback registrado ainda.", styles['Normal']))
        
        # Gerar PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_personal_stats_excel(self, user_id: int) -> bytes:
        """
        Gera relatório Excel com estatísticas pessoais do usuário.
        
        Args:
            user_id: ID do usuário
        
        Returns:
            Bytes do arquivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estatísticas Pessoais"
        
        # Estilos
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="1e40af")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = "Relatório de Estatísticas Pessoais"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        # Informações do usuário
        user = self.db.get_user_by_id(user_id)
        ws[f'A{row}'] = "Usuário:"
        ws[f'B{row}'] = user['full_name'] or user['username']
        row += 1
        ws[f'A{row}'] = "Nível de Experiência:"
        ws[f'B{row}'] = user.get('experience_level', 'N/A')
        row += 1
        ws[f'A{row}'] = "Data do Relatório:"
        ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        row += 2
        
        # Estatísticas principais
        user_stats = self.db.get_user_learning_stats(user_id)
        if user_stats:
            ws[f'A{row}'] = "Métrica"
            ws[f'B{row}'] = "Valor"
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].fill = header_fill
            ws[f'B{row}'].font = header_font
            row += 1
            
            stats = [
                ('Total de Feedbacks', str(user_stats.get('total_feedbacks', 0))),
                ('Taxa de Sucesso', f"{user_stats.get('success_rate', 0) * 100:.1f}%"),
                ('Avaliação Média', f"{user_stats.get('avg_rating', 0):.2f}/5.0"),
                ('Tempo Médio de Execução', f"{user_stats.get('avg_execution_time', 0):.1f}s"),
            ]
            
            for metric, value in stats:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1
            
            row += 1
            
            # Módulos preferidos
            preferred_modules = user_stats.get('preferred_modules', '')
            if preferred_modules:
                ws[f'A{row}'] = "Módulos Preferidos:"
                ws[f'B{row}'] = preferred_modules
                row += 1
        
        # Informações do modelo
        cursor = self.db.conn.cursor()
        cursor.execute("""
            SELECT training_samples, last_trained FROM user_models WHERE user_id = ?
        """, (user_id,))
        model_row = cursor.fetchone()
        
        if model_row:
            row += 1
            ws[f'A{row}'] = "Modelo Personalizado:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = "Amostras de Treinamento:"
            ws[f'B{row}'] = model_row['training_samples'] or 0
            row += 1
            if model_row['last_trained']:
                ws[f'A{row}'] = "Último Treinamento:"
                ws[f'B{row}'] = model_row['last_trained']
        
        # Histórico de feedbacks
        row += 2
        ws[f'A{row}'] = "Histórico de Feedbacks"
        ws[f'A{row}'].font = title_font
        row += 1
        
        feedbacks = self.db.get_user_feedbacks(user_id, limit=100)
        if feedbacks:
            headers = ['Data', 'Teste', 'Sucesso', 'Avaliação', 'Tempo (s)', 'Seguindo Recomendação']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row += 1
            for fb in feedbacks:
                ws.cell(row=row, column=1).value = fb['executed_at'][:10]
                ws.cell(row=row, column=2).value = fb['test_case_id']
                ws.cell(row=row, column=3).value = 'Sim' if fb['success'] else 'Não'
                ws.cell(row=row, column=4).value = fb['tester_rating'] if fb['tester_rating'] else 'N/A'
                ws.cell(row=row, column=5).value = round(fb['actual_execution_time'], 1)
                ws.cell(row=row, column=6).value = 'Sim' if fb['followed_recommendation'] else 'Não'
                
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_comparative_report_excel(self, user_ids: List[int]) -> bytes:
        """
        Gera relatório comparativo entre múltiplos usuários.
        
        Args:
            user_ids: Lista de IDs de usuários
        
        Returns:
            Bytes do arquivo Excel
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparativo de Usuários"
        
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Cabeçalho
        headers = ['Usuário', 'Feedbacks', 'Taxa Sucesso', 'Avaliação Média', 'Tempo Médio', 'Amostras Treino']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Dados de cada usuário
        for user_id in user_ids:
            user = self.db.get_user_by_id(user_id)
            if not user:
                continue
            
            # Buscar estatísticas do usuário
            cursor = self.db.conn.cursor()
            cursor.execute("""
                SELECT * FROM user_learning_stats WHERE user_id = ?
            """, (user_id,))
            row = cursor.fetchone()
            user_stats = dict(row) if row else None
            
            cursor = self.db.conn.cursor()
            cursor.execute("""
                SELECT training_samples FROM user_models WHERE user_id = ?
            """, (user_id,))
            model_row = cursor.fetchone()
            training_samples = model_row['training_samples'] if model_row else 0
            
            ws.cell(row=row, column=1).value = user['full_name'] or user['username']
            ws.cell(row=row, column=2).value = user_stats.get('total_feedbacks', 0) if user_stats else 0
            ws.cell(row=row, column=3).value = f"{user_stats.get('success_rate', 0) * 100:.1f}%" if user_stats else "0%"
            ws.cell(row=row, column=4).value = f"{user_stats.get('avg_rating', 0):.2f}" if user_stats else "0.00"
            ws.cell(row=row, column=5).value = f"{user_stats.get('avg_execution_time', 0):.1f}s" if user_stats else "0.0s"
            ws.cell(row=row, column=6).value = training_samples
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # Ajustar largura das colunas
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def get_model_evolution_data(self, user_id: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Retorna dados de evolução do modelo ao longo do tempo.
        
        Args:
            user_id: ID do usuário (None para modelo global)
        
        Returns:
            Lista de dicionários com dados de evolução
        """
        cursor = self.db.conn.cursor()
        
        if user_id:
            # Evolução do modelo personalizado
            cursor.execute("""
                SELECT 
                    DATE(executed_at) as date,
                    COUNT(*) as feedbacks,
                    AVG(CASE WHEN success = 1 THEN 1.0 ELSE 0.0 END) as success_rate,
                    AVG(tester_rating) as avg_rating,
                    AVG(actual_execution_time) as avg_time
                FROM feedbacks
                WHERE tester_id = ?
                GROUP BY DATE(executed_at)
                ORDER BY date
            """, (user_id,))
        else:
            # Evolução do modelo global
            cursor.execute("""
                SELECT 
                    DATE(executed_at) as date,
                    COUNT(*) as feedbacks,
                    AVG(CASE WHEN success = 1 THEN 1.0 ELSE 0.0 END) as success_rate,
                    AVG(tester_rating) as avg_rating,
                    AVG(actual_execution_time) as avg_time
                FROM feedbacks
                GROUP BY DATE(executed_at)
                ORDER BY date
            """)
        
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
